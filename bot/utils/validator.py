import datetime
import re
from datetime import timedelta
import pandas as pd
import os
import django
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "config.settings")
django.setup()

from django.db.models.functions import ExtractYear
from django.utils import timezone
from django.db.models import Count
from apps.users.models import TelegramUser
from apps.location.models import Region, Area, Village


def validate_uzbek_phone_number(phone_number):
    pattern = re.compile(r'^998\d{9}$')
    qpattern = re.compile(r'^\+998\d{9}$')

    return bool(pattern.match(phone_number) or qpattern.match(phone_number))


def validate_full_name(full_name: str):
    if full_name.startswith("Ro'yxatdan o'tish"):
        return False
    if len(full_name.split(" ")) > 2:
        return True
    else:
        return False


def validate_date_of_birth(year):
    try:
        if (int(datetime.datetime.now().year) - int(year)) > 100 or (int(datetime.datetime.now().year) - int(year)) < 0:
            return False
        return True
    except ValueError as e:
        return False


def strip_tags(html):
    replaced_html = re.sub(r'<(/p).*?>|<(p).*?>|<(br).*?>', '', html)
    replaced_html = replaced_html.replace('<em>', '<i>').replace('</em>', '</i>')
    replaced_html = replaced_html.replace('<strong>', '<b>').replace('</strong>', '</b>')
    return replaced_html


def custom_title_case(s: str):
    result = s.split(maxsplit=2)
    name = ""
    for i in result:
        i = i[0].upper() + i[1:]
        name += i + " "
    return name


def format_testing_time(testing_time):
    testing_time_timedelta = timedelta(
        hours=testing_time.hour,
        minutes=testing_time.minute,
        seconds=testing_time.second
    )

    days, seconds = divmod(testing_time_timedelta.total_seconds(), 86400)
    hours, remainder = divmod(seconds, 3600)
    minutes, seconds = divmod(remainder, 60)

    formatted_time = ""
    if days:
        formatted_time += f"{int(days)} kun "
    if hours:
        formatted_time += f"{int(hours)} soat "
    if minutes:
        formatted_time += f"{int(minutes)} daqiqa "
    if seconds:
        formatted_time += f"{int(seconds)} sekunda"

    return formatted_time.strip()


def correct_answers_counter(data):
    user_answers = data.get("user_answers", {})
    questions = data.get("questions", [])

    correct_answers_count = 0

    for question in questions:
        question_id = question.get("question_id")
        correct_option = question.get("correct_option")
        user_answer = user_answers.get(question_id)

        if user_answer and user_answer == correct_option:
            correct_answers_count += 1

    return correct_answers_count


def excel_upload(data):
    # Create a DataFrame
    df = pd.DataFrame(data)

    # Specify the path where you want to save the Excel file
    output_path = "user_data.xlsx"

    # Write the DataFrame to an Excel file
    df.to_excel(output_path, index=False)

    return output_path


async def generate_statistics():
    current_time = timezone.now().strftime('%d.%m.%Y. Soat %H:%M:%S holatiga')

    total_users = await TelegramUser.objects.acount()

    regions = Region.objects.annotate(user_count=Count('region__area__user_village')).order_by("name")
    regions = [region async for region in regions]

    areas = Area.objects.annotate(user_count=Count('area__user_village')).order_by("name")
    areas = [area async for area in areas]

    result = f"{current_time}\nJami ro'yxatdan o'tganlar: {total_users}\n\n"

    region_stats = []
    for region in regions:
        region_user_count = await TelegramUser.objects.filter(village__area__region=region).acount()
        region_stats.append(f"{region.name} -- {region_user_count}")

    area_stats = []
    for area in areas:
        villages = Village.objects.filter(area=area)
        area_user_count = await TelegramUser.objects.filter(village__in=villages).acount()
        min_required_users = 30 * await villages.acount()
        area_stats.append(f"{area.name} -- {area_user_count}/{min_required_users}")

    current_year = datetime.datetime.now().year
    user_counts = TelegramUser.objects.values('age').annotate(count=Count('id')).filter(count__gt=0).order_by("age")
    user_counts = [user_count async for user_count in user_counts]

    age_counts = {}
    for entry in user_counts:
        birth_year = int(entry['age'])
        age = birth_year
        if age in age_counts:
            age_counts[age] += entry['count']
        else:
            age_counts[age] = entry['count']

    result_string = '\n    '.join([f"{age} yil -- {count}" for age, count in age_counts.items()])

    result += "Hududlar kesimida:\n    " + "\n    ".join(
        region_stats) + "\n\nTuman/shaharlar kesimida:\n    " + "\n    ".join(
        area_stats) + "\n\nYosh kategoriyasi bo'yicha:\n    " + result_string

    return result


async def write_data_to_excel(file_path):
    # Создаем новую книгу Excel
    wb = Workbook()

    # Получаем активный лист
    ws = wb.active

    # Заголовки для столбцов
    ws.append(['Hududlar', 'Tuman/shaharlar', 'Foydalanuvchilar soni', "Mahallarlar soni"])

    # Получаем все регионы
    regions = Region.objects.order_by("name").all()
    regions = [region async for region in regions]

    # Для каждого региона
    for region in regions:
        # Создаем новый лист для региона
        ws_region = wb.create_sheet(title=region.name)

        # Заголовок для листа региона
        ws_region.append(['Tuman/shaharlar', 'Foydalanuvchilar soni', 'Mahallarlar soni'])

        # Получаем все районы в текущем регионе
        areas = region.region.order_by("name").all()
        areas = [area async for area in areas]


        # Для каждого района
        for area in areas:
            villages = Village.objects.filter(area=area).order_by("name")
            villages = [village async for village in villages]
            area_user_count = await TelegramUser.objects.filter(village__in=villages).acount()
            village_count = await area.area.acount() * 30

            # Записываем данные в активный лист
            ws.append([region.name, area.name, f"{area_user_count}", f"{village_count}"])

            # Записываем данные в лист региона
            ws_region.append([area.name, f"{area_user_count}", f"{village_count}"])
        # Автоматически исправляем ширину столбцов на листе региона
        for col in ws_region.columns:
            max_length = 0
            column = col[0].column  # Получаем буквенное представление номера столбца
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws_region.column_dimensions[get_column_letter(column)].width = adjusted_width

    # Автоматически исправляем ширину столбцов на активном листе
    for col in ws.columns:
        max_length = 0
        column = col[0].column  # Получаем буквенное представление номера столбца
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[get_column_letter(column)].width = adjusted_width
    # Сохраняем книгу в файл
    wb.save(file_path)
    return file_path

